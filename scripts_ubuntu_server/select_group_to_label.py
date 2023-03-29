import os
import sys

CURRENT_ROOT = os.path.dirname(os.path.abspath(sys.argv[0]))
ROOT = os.path.join(CURRENT_ROOT, '../')
sys.path.append(ROOT)

from openpyxl import load_workbook
import numpy as np
import pandas as pd

from utils import *

def read_workbook(wb_path, sheetIndex=0):
    # open file
    wb = load_workbook(wb_path, data_only=True)
    worksheets_name_list = wb.sheetnames
    ws = wb[worksheets_name_list[sheetIndex]]
    data = ws.values
    cols = next(data)
    data = list(data)
    df = pd.DataFrame(data, columns=cols)
    return df

def add_normalmode_group_to_label(info_path, dataset_v2_path, output_path, label_per_frames=None):

    if not os.path.exists(output_path):
        os.mkdir(output_path)
        log_GREEN('Create {}'.format(output_path))

    df_info = read_workbook(info_path)
    assert (label_per_frames is not None) or ('num_frames_per_label' in df_info.columns.values.tolist())

    keys_dict = dict()
    for key in df_info.columns.values.tolist():
        keys_dict[key.split('\n')[0]] = key

    num_groups_label = df_info[keys_dict['need_to_label']].values.sum()
    log_BLUE('Need to label {} groups'.format(num_groups_label))

    # calculate number of frame to label
    df_info['num_frames'] = df_info[keys_dict['group_name']].str.split('_', expand=True)[3].str.replace('frames', '').astype('int')
    if label_per_frames is not None:
        df_info['idx_frame_label'] = [np.arange(0, df_info['num_frames'][i], label_per_frames) for i in range(df_info.shape[0])]
    else:
        df_info['idx_frame_label'] = [np.arange(0, df_info['num_frames'][i], df_info['num_frames_per_label'][i]) for i in range(df_info.shape[0])]
    df_info['num_frames_to_label'] = [len(df_info['idx_frame_label'][i]) for i in range(df_info.shape[0])]


    num_frames_to_label = df_info[df_info[keys_dict['need_to_label']]]['num_frames_to_label'].values.sum()
    log_BLUE('Need to label {} frames'.format(num_frames_to_label))

    cnt_groups_label = 0
    for i in range(df_info.shape[0]):
        if not df_info[keys_dict['need_to_label']].iloc[i]:
            log_YELLOW('Skip {} by manual selection'.format(df_info[keys_dict['group_name']].iloc[i]))
            continue

        cnt_groups_label += 1
        cnt_frames_label = 0
        group_folder = os.path.join(dataset_v2_path, df_info[keys_dict['group_name']].iloc[i])
        for idx_frame in df_info['idx_frame_label'].iloc[i]:
            log('='*100)

            cnt_frames_label += 1
            num_frames_in_group = df_info['num_frames'].iloc[i]
            num_frames_label = df_info['idx_frame_label'].iloc[i].shape[0]
            frame_foldername = 'frame{:>04d}'.format(idx_frame)
            frame_folder = os.path.join(group_folder, frame_foldername)

            # build link
            src_path = frame_folder
            dst_path = os.path.join(output_path, df_info[keys_dict['group_name']].iloc[i], frame_foldername)
            if not os.path.exists(os.path.dirname(dst_path)):
                os.mkdir(os.path.dirname(dst_path))
                log_GREEN('Create {}'.format(os.path.dirname(dst_path)))
            try:
                os.remove(dst_path)
                log_YELLOW('Remove {}'.format(dst_path))
            except:
                pass
            os.symlink(src_path, dst_path)
            log_GREEN('Link {}'.format(dst_path))
            log_BLUE(
                'Group in ourDataset_v2: {}/{}, Group to label: {}/{}, Idx_frame in Group: {}/{}, Frame to label: {}/{}'.format(
                    i+1, df_info.shape[0], cnt_groups_label, num_groups_label,
                    idx_frame, num_frames_in_group-1, cnt_frames_label, num_frames_label
                )
            )

def add_mixmode_group_to_label(dataset_v2_path, output_path):

    if not os.path.exists(output_path):
        os.mkdir(output_path)
        log_GREEN('Create {}'.format(output_path))

    group_foldernames = os.listdir(dataset_v2_path)
    group_foldernames.sort()
    for i in range(len(group_foldernames)):
        group_foldername = group_foldernames[i]
        if 'mixmode' in group_foldername:
            log('=' * 100)

            src_path = os.path.join(dataset_v2_path, group_foldername)
            dst_path = os.path.join(output_path, group_foldername)
            try:
                os.remove(dst_path)
                log_YELLOW('Remove {}'.format(dst_path))
            except:
                pass
            os.symlink(src_path, dst_path)

            log_GREEN('Link {}'.format(group_foldername))

def add_normalmode_per_5frames():
    info_path = '/mnt/Dataset/ourDataset_v2_group_infomation.xlsx'
    dataset_v2_path = '/mnt/Dataset/ourDataset_v2'
    output_path = '/mnt/Dataset/ourDataset_v2_label'
    add_normalmode_group_to_label(info_path, dataset_v2_path, output_path, label_per_frames=5)

def add_mixmode():
    dataset_v2_path = '/mnt/Dataset/ourDataset_v2'
    output_path = '/mnt/Dataset/ourDataset_v2_label'
    add_mixmode_group_to_label(dataset_v2_path, output_path)

def add_group_for_tracking():
    info_path = '/mnt/Dataset/ourDataset_v2_group_infomation_for_tracking.xlsx'
    dataset_v2_path = '/mnt/Dataset/ourDataset_v2'
    output_path = '/mnt/Dataset/ourDataset_v2_label'
    add_normalmode_group_to_label(info_path, dataset_v2_path, output_path)

def main():
    add_normalmode_per_5frames()

    add_mixmode()

    add_group_for_tracking()


if __name__ == '__main__':
    main()

