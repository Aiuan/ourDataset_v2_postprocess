import os
import sys
import glob
import shutil

CURRENT_ROOT = os.path.dirname(os.path.abspath(sys.argv[0]))
ROOT = os.path.join(CURRENT_ROOT, '../')
sys.path.append(ROOT)

import numpy as np
from openpyxl import load_workbook
import pandas as pd

from dataset_v2 import log, log_GREEN, load_json, save_dict_as_json, log_YELLOW


def check_to_create_path(path):
    if not os.path.exists(path):
        os.mkdir(path)
        log_GREEN('Create {}'.format(path))

def read_workbook(wb_path, sheetIndex=0):
    # open file
    wb = load_workbook(wb_path, data_only=True)
    worksheets_name_list = wb.sheetnames
    ws = wb[worksheets_name_list[sheetIndex]]
    data = ws.values
    cols = next(data)
    data = list(data)
    df = pd.DataFrame(data, columns=cols)
    return df

def main():
    root_dataset = '/mnt/Dataset/ourDataset_v2'
    xlsx_path = '/mnt/Dataset/data_for_SensorCalibration.xlsx'
    root_output = '/mnt/Dataset/data_for_SensorCalibration'
    cameras = ['IRayCamera', 'LeopardCamera0', 'LeopardCamera1']
    pcd_sensor = 'VelodyneLidar'

    # read xlsx
    df = read_workbook(xlsx_path)

    check_to_create_path(root_output)

    for i in range(df.shape[0]):
        log('=' * 100)

        groupname = df['group'].iloc[i]
        idx_frame = df['idx_frame'].iloc[i]
        framename = 'frame{:>04d}'.format(idx_frame)

        log('Process {}/{} {}:{}'.format(i + 1, df.shape[0], groupname, framename))

        frame_folder_path = os.path.join(root_dataset, groupname, framename)

        root_output_scene = os.path.join(root_output, 'scene{}'.format(i))
        check_to_create_path(root_output_scene)

        for camera in cameras:
            image_src_path = glob.glob(os.path.join(frame_folder_path, camera, '*.png'))[0]
            image_dst_path = os.path.join(root_output_scene, '{}.png'.format(camera))
            shutil.copy(image_src_path, image_dst_path)
            log_GREEN('Copy to {}'.format(image_dst_path))

        pcd_src_path = glob.glob(os.path.join(frame_folder_path, pcd_sensor, '*.pcd'))[0]
        pcd_dst_path = os.path.join(root_output_scene, '{}.pcd'.format(pcd_sensor))
        shutil.copy(pcd_src_path, pcd_dst_path)
        log_GREEN('Copy to {}'.format(pcd_dst_path))


if __name__ == '__main__':
    main()
